import io
from decimal import Decimal
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


STATUS_LABELS = {"pendente": "Pendente", "pago": "Pago", "vencido": "Vencido"}
HEADER_FILL = "1B4F8C"
SUBHEADER_FILL = "2E86AB"


def gerar_recibo_pdf(parcela, empresa=None):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    blue = colors.HexColor("#1B4F8C")

    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=18, textColor=blue, spaceAfter=4)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, textColor=colors.grey, spaceAfter=12)
    label_style = ParagraphStyle("label", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    value_style = ParagraphStyle("value", parent=styles["Normal"], fontSize=11, fontName="Helvetica-Bold")
    footer_style = ParagraphStyle("footer", parent=styles["Normal"], fontSize=8, textColor=colors.grey, alignment=1)

    elements = []

    nome_empresa = empresa.nome if empresa else "Gestão de Consórcios"
    elements.append(Paragraph(nome_empresa, title_style))
    if empresa and empresa.cnpj:
        elements.append(Paragraph(f"CNPJ: {empresa.cnpj}", sub_style))
    elements.append(Paragraph("RECIBO DE PAGAMENTO DE COMISSÃO", ParagraphStyle("rec", parent=styles["Heading2"], fontSize=13, textColor=blue, spaceAfter=8)))
    elements.append(HRFlowable(width="100%", thickness=1, color=blue, spaceAfter=12))

    venda = parcela.venda
    vendedor = venda.vendedor
    cliente = venda.cliente

    dados = [
        ["Contrato", venda.numero_contrato],
        ["Data da Venda", venda.data_venda.strftime("%d/%m/%Y")],
        ["Vendedor", vendedor.nome],
        ["Coordenador", vendedor.coordenador.nome],
        ["Cliente", cliente.nome if cliente else "—"],
        ["Consórcio", venda.consorcio.nome],
        ["Valor do Bem", f"R$ {venda.valor_bem:,.2f}"],
        ["Parcela", f"{parcela.numero_parcela}"],
        ["Vencimento", parcela.data_vencimento.strftime("%d/%m/%Y")],
        ["Data de Pagamento", parcela.data_pagamento.strftime("%d/%m/%Y") if parcela.data_pagamento else "—"],
        ["Valor da Comissão", f"R$ {parcela.valor:,.2f}"],
        ["Status", "PAGO"],
    ]

    table_data = [[Paragraph(f"<b>{k}</b>", label_style), Paragraph(str(v), value_style)] for k, v in dados]
    t = Table(table_data, colWidths=[6*cm, 11*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F0F4F8")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWHEIGHT", (0, 0), (-1, -1), 22),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("BACKGROUND", (0, -2), (-1, -1), colors.HexColor("#D4EDDA")),
        ("FONTNAME", (0, -2), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 1*cm))

    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey, spaceAfter=8))
    elements.append(Paragraph(f"Gerado em: {date.today().strftime('%d/%m/%Y')}", footer_style))
    if empresa and empresa.texto_recibo:
        elements.append(Paragraph(empresa.texto_recibo, footer_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def gerar_excel_comissoes(parcelas_qs, titulo="Relatório de Comissões"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comissões"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:I1")
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=14, color=HEADER_FILL)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Gerado em: {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].alignment = Alignment(horizontal="right")
    ws.row_dimensions[2].height = 18

    headers = ["Contrato", "Vendedor", "Coordenador", "Parcela", "Vencimento", "Valor (R$)", "Status", "Pagamento"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[4].height = 22

    col_widths = [18, 25, 25, 10, 15, 15, 12, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

    status_colors = {"pendente": "FFF3CD", "pago": "D4EDDA", "vencido": "F8D7DA"}
    total = Decimal("0")

    for row_num, p in enumerate(parcelas_qs, 5):
        row_fill = PatternFill(
            start_color=status_colors.get(p.status, "FFFFFF"),
            end_color=status_colors.get(p.status, "FFFFFF"),
            fill_type="solid"
        )
        row_data = [
            p.venda.numero_contrato,
            p.venda.vendedor.nome,
            p.venda.vendedor.coordenador.nome,
            p.numero_parcela,
            p.data_vencimento.strftime("%d/%m/%Y"),
            float(p.valor),
            STATUS_LABELS.get(p.status, p.status),
            p.data_pagamento.strftime("%d/%m/%Y") if p.data_pagamento else "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = border
            cell.fill = row_fill
            if col == 6:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

        total += p.valor

    last_row = parcelas_qs.count() + 5
    ws.cell(row=last_row, column=5, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=last_row, column=6, value=float(total))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'
    total_cell.alignment = Alignment(horizontal="right")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def gerar_pdf_comissoes(parcelas_qs, titulo="Relatório de Comissões"):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#1B4F8C"))
    elements.append(Paragraph(titulo, title_style))
    elements.append(Paragraph(f"Gerado em: {date.today().strftime('%d/%m/%Y')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5*cm))

    headers = ["Contrato", "Vendedor", "Coordenador", "Parc.", "Vencimento", "Valor (R$)", "Status", "Pagamento"]
    data = [headers]
    total = Decimal("0")

    for p in parcelas_qs:
        data.append([
            p.venda.numero_contrato,
            p.venda.vendedor.nome[:20],
            p.venda.vendedor.coordenador.nome[:20],
            str(p.numero_parcela),
            p.data_vencimento.strftime("%d/%m/%Y"),
            f"R$ {p.valor:,.2f}",
            STATUS_LABELS.get(p.status, p.status),
            p.data_pagamento.strftime("%d/%m/%Y") if p.data_pagamento else "-",
        ])
        total += p.valor

    data.append(["", "", "", "", "TOTAL", f"R$ {total:,.2f}", "", ""])

    col_widths = [3.5*cm, 5*cm, 5*cm, 1.5*cm, 3*cm, 3.5*cm, 2.5*cm, 3*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    status_row_colors = {"pendente": colors.HexColor("#FFF3CD"), "pago": colors.HexColor("#D4EDDA"), "vencido": colors.HexColor("#F8D7DA")}
    row_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4F8C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWHEIGHT", (0, 0), (-1, -1), 18),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]

    for i, p in enumerate(parcelas_qs, 1):
        bg = status_row_colors.get(p.status, colors.white)
        row_commands.append(("BACKGROUND", (0, i), (-1, i), bg))

    table.setStyle(TableStyle(row_commands))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer
